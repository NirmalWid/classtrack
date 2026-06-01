import pyrebase

from firebase_config import firebaseConfig

from sms_service import send_sms

from datetime import datetime

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    send_file,
    Response,
    jsonify
)

import firebase_admin
from firebase_admin import credentials, firestore

import qrcode
import io

from datetime import datetime

from dotenv import load_dotenv

from io import BytesIO

from flask import send_file

from openpyxl import Workbook

import os

import json

# =========================================
# LOAD ENV
# =========================================

load_dotenv()


# =========================================
# FLASK SETUP
# =========================================

app = Flask(__name__)

app.secret_key = os.getenv("SECRET_KEY")


# =========================================
# FIREBASE SETUP
# =========================================

firebase_credentials = json.loads(os.environ["FIREBASE_CREDENTIALS"])

cred = credentials.Certificate(firebase_credentials)

firebase_admin.initialize_app(cred)

db = firestore.client()

firebase = pyrebase.initialize_app(
    firebaseConfig
)

auth = firebase.auth()

# =========================================
# HELPER FUNCTIONS
# =========================================

def current_date():
    return datetime.now().strftime("%Y-%m-%d")


def current_time():
    return datetime.now().strftime("%H:%M:%S")


def current_month():
    return datetime.now().strftime("%Y-%m")


# =========================================
# LOGIN
# =========================================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    error = None

    if request.method == "POST":

        email = request.form["email"]

        password = request.form["password"]

        try:

            user = auth.sign_in_with_email_and_password(

                email,
                password
            )

            session["user"] = email

            return redirect(
                url_for("dashboard")
            )

        except:

            error = "Invalid login"

    return render_template(
        "login.html",
        error=error
    )


# =========================================
# LOGOUT
# =========================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect(
        url_for("login")
    )


# =========================================
# LOGIN REQUIRED
# =========================================

from functools import wraps

def login_required(f):

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if "user" not in session:

            return redirect(
                url_for("login")
            )

        return f(*args, **kwargs)

    return decorated_function

# =========================================
# DASHBOARD
# =========================================

@app.route("/")
@login_required
def dashboard():

    return render_template("dashboard.html")


# =========================================
# START CLASS + ABSENT SMS
# =========================================

@app.route("/start-class/<class_id>")
@login_required
def start_class(class_id):

    today = datetime.now().strftime("%Y-%m-%d")

    class_doc = db.collection("classes").document(class_id).get()

    if not class_doc.exists:
        return "Class not found"

    class_data = class_doc.to_dict()
    class_display = f"{class_data.get('grade','')} - {class_data.get('class_type','')}"

    run_id = f"{class_id}_{today}"
    run_ref = db.collection("absent_runs").document(run_id)

    # ===============================
    # IF ALREADY PROCESSED
    # ===============================
    if run_ref.get().exists:

        logs_ref = db.collection("absent_sms_logs")\
            .where("class_id", "==", class_id)\
            .where("date", "==", today)\
            .stream()

        absent_students = [doc.to_dict() for doc in logs_ref]

        return render_template(
            "start_class_result.html",
            class_display=class_display,
            date=today,
            absent_students=absent_students,
            already_sent=True
        )

    # mark run
    run_ref.set({
        "class_id": class_id,
        "date": today,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

    students_ref = db.collection("students").where(
        "class_id", "==", class_id
    ).stream()

    students = [doc.to_dict() for doc in students_ref]

    absent_students = []

    for student in students:

        student_id = student["student_id"]
        attendance_id = f"{student_id}_{today}"

        attendance_doc = db.collection("attendance").document(attendance_id).get()

        if not attendance_doc.exists:

            sms_id = f"{student_id}_{today}"
            sms_ref = db.collection("absent_sms_logs").document(sms_id)

            if sms_ref.get().exists:
                continue

            sms_message = (
                f"Dear Parent, {student['student_name']} "
                f"was absent for {class_display} class today."
            )

            try:
                send_sms(student["parent_phone"], sms_message)

                sms_ref.set({
                    "student_id": student_id,
                    "student_name": student["student_name"],
                    "class_id": class_id,
                    "class_display": class_display,
                    "date": today,
                    "parent_phone": student["parent_phone"],
                    "message": sms_message
                })

                absent_students.append({
                    "student_name": student["student_name"],
                    "parent_phone": student["parent_phone"]
                })

            except Exception as e:
                print("SMS ERROR:", e)

    return render_template(
        "start_class_result.html",
        class_display=class_display,
        date=today,
        absent_students=absent_students,
        already_sent=False
    )

# =========================================
# QR GENERATION
# =========================================

@app.route("/qr/<student_id>")
@login_required
def generate_qr(student_id):

    qr = qrcode.make(student_id)

    img_io = io.BytesIO()

    qr.save(img_io, "PNG")

    img_io.seek(0)

    return send_file(
        img_io,
        mimetype="image/png"
    )
    
# =========================================
# Add Classes
# =========================================    
    
@app.route("/classes", methods=["GET", "POST"])
@login_required
def classes():

    if request.method == "POST":

        grade = request.form["grade"]

        class_type = request.form["class_type"]

        class_day = request.form["class_day"]

        start_time = request.form["start_time"]

        end_time = request.form["end_time"]

        monthly_fee = request.form["monthly_fee"]

        # ==========================
        # DUPLICATE CHECK
        # ==========================

        existing_class = db.collection(
            "classes"
        ).where(
            "grade", "==", grade
        ).where(
            "class_type", "==", class_type
        ).where(
            "class_day", "==", class_day
        ).where(
            "start_time", "==", start_time
        ).stream()

        if len(list(existing_class)) > 0:

            classes_ref = db.collection(
                "classes"
            ).stream()

            classes_list = [
                doc.to_dict()
                for doc in classes_ref
            ]

            return render_template(
                "classes.html",
                classes=classes_list,
                error="This class already exists."
            )

        # ==========================
        # CREATE CLASS
        # ==========================

        class_ref = db.collection(
            "classes"
        ).document()

        class_data = {

            "class_id": class_ref.id,

            "grade": grade,

            "class_type": class_type,

            "class_display": (
                f"Grade {grade} - {class_type}"
            ),

            "class_day": class_day,

            "start_time": start_time,

            "end_time": end_time,

            "monthly_fee": int(monthly_fee),

            "active": True,

            "created_at": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        }

        class_ref.set(class_data)

        return redirect(
            url_for("classes")
        )

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes_list = [
        doc.to_dict()
        for doc in classes_ref
    ]

    return render_template(
        "classes.html",
        classes=classes_list
    )

# =========================================
# Delete clz
# =========================================

@app.route("/delete-class/<class_id>")
@login_required
def delete_class(class_id):

    db.collection("classes").document(class_id).delete()

    return redirect(url_for("classes"))


# =========================================
# ADD STUDENT
# =========================================

@app.route(
    "/add-student",
    methods=["GET", "POST"]
)
@login_required
def add_student():

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [
        doc.to_dict()
        for doc in classes_ref
    ]

    if request.method == "POST":

        student_name = request.form[
            "student_name"
        ].strip()

        parent_name = request.form[
            "parent_name"
        ].strip()

        parent_phone = request.form[
            "parent_phone"
        ].strip()

        class_id = request.form[
            "class_id"
        ]

        # ==========================
        # DUPLICATE CHECK
        # ==========================

        existing_student = db.collection(
            "students"
        ).where(
            "student_name",
            "==",
            student_name
        ).where(
            "parent_phone",
            "==",
            parent_phone
        ).stream()

        if len(list(existing_student)) > 0:

            return render_template(
                "add_student.html",
                classes=classes,
                error="Student already exists."
            )

        # ==========================
        # GET CLASS
        # ==========================

        class_doc = db.collection(
            "classes"
        ).document(class_id).get()

        if not class_doc.exists:

            return "Class not found"

        class_data = class_doc.to_dict()

        class_display = (

            f"Grade {class_data['grade']}"

            f" - "

            f"{class_data['class_type']}"
        )

        # ==========================
        # CREATE STUDENT
        # ==========================

        student_ref = db.collection(
            "students"
        ).document()

        student_data = {

            "student_id": student_ref.id,

            "student_name": student_name,

            "parent_name": parent_name,

            "parent_phone": parent_phone,

            "class_id": class_id,

            # OLD SUPPORT
            "class_name": class_display,

            # NEW STANDARD
            "class_display": class_display,

            "active": True,

            "created_at": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        }

        student_ref.set(student_data)

        return redirect(
            url_for("students")
        )

    return render_template(
        "add_student.html",
        classes=classes
    )

# =========================================
# EDIT STUDENT
# =========================================

@app.route(
    "/edit-student/<student_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_student(student_id):

    student_ref = db.collection(
        "students"
    ).document(student_id)

    student_doc = student_ref.get()

    if not student_doc.exists:

        return "Student not found"

    student = student_doc.to_dict()

    # GET CLASSES

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [

        doc.to_dict()

        for doc in classes_ref
    ]

    if request.method == "POST":

        class_id = request.form[
            "class_id"
        ]

        class_doc = db.collection(
            "classes"
        ).document(class_id).get()

        if not class_doc.exists:

            return "Class not found"

        class_data = class_doc.to_dict()

        class_display = (

            f"Grade {class_data.get('grade')} "

            f"- "

            f"{class_data.get('class_type')}"
        )

        student_ref.update({

            "student_name": request.form[
                "student_name"
            ],

            "parent_name": request.form[
                "parent_name"
            ],

            "parent_phone": request.form[
                "parent_phone"
            ],

            "class_id": class_id,

            # OLD SUPPORT
            "class_name": class_display,

            # NEW STANDARD
            "class_display": class_display
        })

        return redirect(
            url_for("students")
        )

    return render_template(

        "edit_student.html",

        student=student,

        classes=classes
    )

# =========================================
# delete student
# =========================================

@app.route("/delete-student/<student_id>")
@login_required
def delete_student(student_id):

    # DELETE STUDENT

    db.collection(
        "students"
    ).document(student_id).delete()

    # DELETE ATTENDANCE

    attendance_docs = db.collection(
        "attendance"
    ).where(
        "student_id",
        "==",
        student_id
    ).stream()

    for doc in attendance_docs:

        doc.reference.delete()

    # DELETE FEES

    fee_docs = db.collection(
        "fees"
    ).where(
        "student_id",
        "==",
        student_id
    ).stream()

    for doc in fee_docs:

        doc.reference.delete()

    return redirect(
        url_for("students")
    )


# =========================================
# Students
# =========================================


@app.route("/students")
@login_required
def students():

    search = request.args.get(
        "search",
        ""
    )

    class_filter = request.args.get(
        "class_id",
        ""
    )

    students_ref = db.collection(
        "students"
    ).stream()

    students = [
        doc.to_dict()
        for doc in students_ref
    ]

    # SEARCH FILTER

    if search:

        students = [

            s for s in students

            if search.lower() in
            s.get(
                "student_name",
                ""
            ).lower()

        ]

    # CLASS FILTER

    if class_filter:

        students = [

            s for s in students

            if s.get("class_id") == class_filter

        ]

    # GET CLASSES

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [
        doc.to_dict()
        for doc in classes_ref
    ]

    return render_template(

        "students.html",

        students=students,

        classes=classes,

        search=search,

        class_filter=class_filter
    )


# =========================================
# Download QR
# =========================================

@app.route("/download-qr/<student_id>")
@login_required
def download_qr(student_id):

    student_doc = db.collection(
        "students"
    ).document(student_id).get()

    if not student_doc.exists:
        return "Student not found"

    student = student_doc.to_dict()

    filename = (
        student["student_name"]
        .replace(" ", "_")
        + "_QR.png"
    )

    qr = qrcode.make(student_id)

    img_io = io.BytesIO()

    qr.save(img_io, "PNG")

    img_io.seek(0)

    return send_file(
        img_io,
        mimetype="image/png",
        as_attachment=True,
        download_name=filename
    )

# =========================================
# QR scan
# =========================================

@app.route("/qr/<student_id>")
@login_required
def student_qr(student_id):

    qr = qrcode.make(student_id)

    buffer = BytesIO()

    qr.save(buffer, format="PNG")

    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="image/png"
    )


@app.route("/scan", methods=["GET", "POST"])
@login_required
def scan():

    message = None
    student = None
    fee_status = "Not Paid"
    show_paid_button = True
    scanned = False

    # =====================================
    # GET
    # =====================================
    if request.method == "GET":

        reset = request.args.get("reset")

        if reset:
            return render_template(
                "scan.html",
                message=None,
                student=None,
                fee_status="Not Paid",
                show_paid_button=True,
                scanned=False
            )

        scanned_student_id = request.args.get("student_id")

        if scanned_student_id:

            student_doc = db.collection("students").document(scanned_student_id).get()

            if student_doc.exists:

                student = student_doc.to_dict()
                scanned = True

                current_month = datetime.now().strftime("%Y-%m")

                fee_doc = db.collection("fees").document(
                    f"{scanned_student_id}_{current_month}"
                ).get()

                if fee_doc.exists:
                    fee_status = fee_doc.to_dict().get("status", "Not Paid")

                show_paid_button = (fee_status != "Paid")

    # =====================================
    # POST
    # =====================================
    if request.method == "POST":

        action = request.form.get("action")
        student_id = request.form.get("student_id")

        if not student_id:
            return redirect(url_for("scan"))

        student_doc = db.collection("students").document(student_id).get()

        if not student_doc.exists:
            return redirect(url_for("scan", message="Invalid QR"))

        student = student_doc.to_dict()
        scanned = True

        current_month = datetime.now().strftime("%Y-%m")

        fee_doc_id = f"{student_id}_{current_month}"
        fee_ref = db.collection("fees").document(fee_doc_id)

        fee_doc = fee_ref.get()

        if fee_doc.exists:
            fee_status = fee_doc.to_dict().get("status", "Not Paid")

        # =====================================
        # PAY
        # =====================================
        if action == "pay":

            fee_ref.set({
                "student_id": student_id,
                "student_name": student["student_name"],
                "class_id": student["class_id"],
                "class_display": student.get("class_display", ""),
                "month": current_month,
                "status": "Paid",
                "paid_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })

            fee_status = "Paid"
            show_paid_button = False

        # =====================================
        # ATTENDANCE
        # =====================================
        if action == "attendance":

            today = datetime.now().strftime("%Y-%m-%d")
            attendance_id = f"{student_id}_{today}"
            attendance_ref = db.collection("attendance").document(attendance_id)

            if attendance_ref.get().exists:
                return redirect(url_for("scan", message="Already marked today", reset=1))

            attendance_ref.set({
                "attendance_id": attendance_id,
                "student_id": student_id,
                "student_name": student["student_name"],
                "class_id": student["class_id"],
                "class_display": student.get("class_display", ""),
                "date": today,
                "time": datetime.now().strftime("%H:%M:%S"),
                "fee_status": fee_status,
                "status": "Present"
            })

            # =====================================
            # SMS LOGIC (FIXED)
            # =====================================
            time_now = datetime.now().strftime("%H:%M")

            base_message = (
                f"Dear Parent, {student['student_name']} attended "
                f"{student.get('class_display','Class')} at {time_now}."
            )

            if fee_status == "Paid":
                sms_message = base_message + " Class fees : Received."
            else:
                sms_message = base_message

            try:
                send_sms(student["parent_phone"], sms_message)
                print("SMS SENT:", sms_message)
            except Exception as e:
                print("SMS ERROR:", e)

            return redirect(url_for(
                "scan",
                message=f"Attendance marked for {student['student_name']}",
                reset=1
            ))

    # =====================================
    # RENDER
    # =====================================
    if request.args.get("message"):
        message = request.args.get("message")

    return render_template(
        "scan.html",
        message=message,
        student=student,
        fee_status=fee_status,
        show_paid_button=show_paid_button,
        scanned=scanned
    )


@app.route("/attendance")
@login_required
def attendance():

    selected_class = request.args.get("class_id")

    today = datetime.now().strftime("%Y-%m-%d")

    # =====================================
    # GET CLASSES
    # =====================================
    classes_ref = db.collection("classes").stream()

    classes = [
        doc.to_dict()
        for doc in classes_ref
    ]

    # =====================================
    # GET STUDENTS
    # =====================================
    students_query = db.collection("students")

    if selected_class:
        students_query = students_query.where(
            "class_id",
            "==",
            selected_class
        )

    students = [
        doc.to_dict()
        for doc in students_query.stream()
    ]

    # =====================================
    # GET TODAY ATTENDANCE
    # =====================================
    attendance_ref = db.collection("attendance").where(
        "date",
        "==",
        today
    ).stream()

    attendance_records = [
        doc.to_dict()
        for doc in attendance_ref
    ]

    # =====================================
    # FILTER BY CLASS (if selected)
    # =====================================
    if selected_class:
        attendance_records = [
            a for a in attendance_records
            if a.get("class_id") == selected_class
        ]

    # =====================================
    # PRESENT STUDENT IDS (SAFE)
    # =====================================
    present_ids = [
        a.get("student_id")
        for a in attendance_records
        if a.get("student_id")
    ]

    # =====================================
    # ABSENT STUDENTS
    # =====================================
    absent_students = [
        s for s in students
        if s.get("student_id") not in present_ids
    ]

    # =====================================
    # COUNTS
    # =====================================
    total_students = len(students)
    total_present = len(attendance_records)
    total_absent = len(absent_students)

    return render_template(
        "attendance.html",

        classes=classes,
        attendance_records=attendance_records,
        absent_students=absent_students,

        total_students=total_students,
        total_present=total_present,
        total_absent=total_absent,

        selected_class=selected_class,
        today=today
    )


# =========================================
# clz fees
# =========================================

@app.route("/fees")
@login_required
def fees():

    selected_class = request.args.get(
        "class_id"
    )

    current_month = datetime.now().strftime(
        "%Y-%m"
    )

    # GET CLASSES

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [

        doc.to_dict()

        for doc in classes_ref
    ]

    # GET STUDENTS

    students_query = db.collection(
        "students"
    )

    if selected_class:

        students_query = students_query.where(
            "class_id",
            "==",
            selected_class
        )

    students_ref = students_query.stream()

    students = [

        doc.to_dict()

        for doc in students_ref
    ]

    # GET FEES

    fees_ref = db.collection(
        "fees"
    ).where(
        "month",
        "==",
        current_month
    ).stream()

    fee_records = [

        doc.to_dict()

        for doc in fees_ref
    ]

    # FILTER FEES

    if selected_class:

        fee_records = [

            f for f in fee_records

            if f.get("class_id")
            == selected_class
        ]

    # PAID IDS

    paid_ids = [

        f["student_id"]

        for f in fee_records

        if f.get("status") == "Paid"
    ]

    # PAID STUDENTS

    paid_students = [

        s for s in students

        if s["student_id"]
        in paid_ids
    ]

    # UNPAID STUDENTS

    unpaid_students = [

        s for s in students

        if s["student_id"]
        not in paid_ids
    ]

    # REVENUE

    total_revenue = 0

    for student in paid_students:

        class_doc = db.collection(
            "classes"
        ).document(
            student["class_id"]
        ).get()

        if class_doc.exists:

            class_data = class_doc.to_dict()

            total_revenue += class_data.get(
                "monthly_fee",
                0
            )

    return render_template(

        "fees.html",

        classes=classes,

        paid_students=paid_students,

        unpaid_students=unpaid_students,

        total_revenue=total_revenue,

        current_month=current_month,

        selected_class=selected_class
    )


# =========================================
# edit clz
# =========================================

@app.route(
    "/edit-class/<class_id>",
    methods=["GET", "POST"]
)
@login_required
def edit_class(class_id):

    class_ref = db.collection("classes").document(class_id)

    class_doc = class_ref.get()

    if not class_doc.exists:

        return "Class not found"

    if request.method == "POST":

        grade = request.form["grade"]

        class_type = request.form["class_type"]

        class_ref.update({

            "grade": grade,

            "class_type": class_type,

            "class_display": (
                f"{grade} - {class_type}"
            ),

            "class_day": request.form["class_day"],

            "start_time": request.form["start_time"],

            "end_time": request.form["end_time"],

            "monthly_fee": int(
                request.form["monthly_fee"]
            )

        })

        return redirect(url_for("classes"))

    class_data = class_doc.to_dict()

    return render_template(
        "edit_class.html",
        class_data=class_data
    )


@app.route("/student/<student_id>")
@login_required
def student_profile(student_id):

    # GET STUDENT

    student_doc = db.collection(
        "students"
    ).document(student_id).get()

    if not student_doc.exists:

        return "Student not found"

    student = student_doc.to_dict()

    # GET ATTENDANCE HISTORY

    attendance_ref = db.collection(
        "attendance"
    ).where(
        "student_id",
        "==",
        student_id
    ).stream()

    attendance_history = [

        doc.to_dict()

        for doc in attendance_ref
    ]

    # SORT NEWEST FIRST

    attendance_history = sorted(

        attendance_history,

        key=lambda x: x.get(
            "date",
            ""
        ),

        reverse=True
    )

    # GET FEE HISTORY

    fees_ref = db.collection(
        "fees"
    ).where(
        "student_id",
        "==",
        student_id
    ).stream()

    fee_history = [

        doc.to_dict()

        for doc in fees_ref
    ]

    fee_history = sorted(

        fee_history,

        key=lambda x: x.get(
            "month",
            ""
        ),

        reverse=True
    )

    # STATS

    total_attendance = len(
        attendance_history
    )

    total_paid_months = len([

        f for f in fee_history

        if f.get("status") == "Paid"
    ])

    return render_template(

        "student_profile.html",

        student=student,

        attendance_history=attendance_history,

        fee_history=fee_history,

        total_attendance=total_attendance,

        total_paid_months=total_paid_months
    )

# =========================================
# reports
# =========================================

@app.route("/reports")
@login_required
def reports():

    report_type = request.args.get(
        "report_type"
    )

    from_date = request.args.get(
        "from_date"
    )

    to_date = request.args.get(
        "to_date"
    )

    month = request.args.get(
        "month"
    )

    class_id = request.args.get(
    "class_id"
    )

    attendance_records = []

    class_students = []

    paid_students = []

    unpaid_students = []

    total_revenue = 0

    # GET CLASSES

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [

        doc.to_dict()

        for doc in classes_ref
    ]

    # ATTENDANCE REPORT

    if report_type == "attendance":

        attendance_ref = db.collection(
            "attendance"
        ).stream()

        for doc in attendance_ref:

            a = doc.to_dict()

            attendance_date = a.get(
                "date"
            )

            if from_date <= attendance_date <= to_date:

                attendance_records.append(a)

    # CLASS STUDENT REPORT

    if report_type == "class_students":

        students_ref = db.collection(
            "students"
        ).where(
            "class_id",
            "==",
            class_id
        ).stream()

        class_students = [

            doc.to_dict()

            for doc in students_ref
        ]            

    # FEES REPORT

    if report_type == "fees":

        students_ref = db.collection(
            "students"
        ).stream()

        students = [

            doc.to_dict()

            for doc in students_ref
        ]

        fees_ref = db.collection(
            "fees"
        ).where(
            "month",
            "==",
            month
        ).stream()

        fees = [

            doc.to_dict()

            for doc in fees_ref
        ]

        paid_ids = [

            f["student_id"]

            for f in fees
        ]

        paid_students = [

            s for s in students

            if s["student_id"]
            in paid_ids
        ]

        unpaid_students = [

            s for s in students

            if s["student_id"]
            not in paid_ids
        ]

        # REVENUE

        for student in paid_students:

            class_doc = db.collection(
                "classes"
            ).document(
                student["class_id"]
            ).get()

            if class_doc.exists:

                class_data = class_doc.to_dict()

                total_revenue += class_data.get(
                    "monthly_fee",
                    0
                )

    return render_template(

        "reports.html",

        class_students=class_students,
        
        class_id=class_id,

        classes=classes,

        report_type=report_type,

        from_date=from_date,

        to_date=to_date,

        month=month,

        attendance_records=attendance_records,

        paid_students=paid_students,

        unpaid_students=unpaid_students,

        total_revenue=total_revenue
    )


@app.route("/export-attendance")
@login_required
def export_attendance():

    from_date = request.args.get(
        "from_date"
    )

    to_date = request.args.get(
        "to_date"
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"

    # HEADERS

    headers = [

        "Date",

        "Student Name",

        "Class",

        "Time",

        "Fee Status"
    ]

    ws.append(headers)

    attendance_ref = db.collection(
        "attendance"
    ).stream()

    for doc in attendance_ref:

        a = doc.to_dict()

        attendance_date = a.get(
            "date"
        )

        if from_date <= attendance_date <= to_date:

            ws.append([

                a.get("date"),

                a.get("student_name"),

                a.get("class_name"),

                a.get("time"),

                a.get("fee_status")
            ])

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            f"attendance_"
            f"{from_date}_"
            f"{to_date}.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )



@app.route("/export-fees")
@login_required
def export_fees():

    month = request.args.get("month")

    wb = Workbook()

    ws = wb.active

    ws.title = "Fees"

    headers = [

        "Student Name",

        "Class",

        "Month",

        "Status",

        "Parent Phone",

        "Paid At"
    ]

    ws.append(headers)

    # ====================================
    # GET ALL STUDENTS
    # ====================================

    students_ref = db.collection(
        "students"
    ).stream()

    students = [

        doc.to_dict()

        for doc in students_ref
    ]

    # ====================================
    # GET FEES FOR MONTH
    # ====================================

    fees_ref = db.collection(
        "fees"
    ).where(
        "month",
        "==",
        month
    ).stream()

    fees = [

        doc.to_dict()

        for doc in fees_ref
    ]

    paid_ids = [

        fee.get("student_id")

        for fee in fees
    ]

    # ====================================
    # PAID STUDENTS
    # ====================================

    for fee in fees:

        class_name = ""

        class_doc = db.collection(
            "classes"
        ).document(
            fee.get("class_id")
        ).get()

        if class_doc.exists:

            class_data = class_doc.to_dict()

            class_name = (

                f"Grade {class_data.get('grade')} "

                f"- "

                f"{class_data.get('class_type')}"
            )

        ws.append([

            fee.get(
                "student_name",
                ""
            ),

            class_name,

            fee.get(
                "month",
                ""
            ),

            "Paid",

            fee.get(
                "parent_phone",
                ""
            ),

            fee.get(
                "paid_at",
                ""
            )
        ])

    # ====================================
    # UNPAID STUDENTS
    # ====================================

    for student in students:

        if student.get("student_id") not in paid_ids:

            class_name = ""

            class_doc = db.collection(
                "classes"
            ).document(
                student.get("class_id")
            ).get()

            if class_doc.exists:

                class_data = class_doc.to_dict()

                class_name = (

                    f"Grade {class_data.get('grade')} "

                    f"- "

                    f"{class_data.get('class_type')}"
                )

            ws.append([

                student.get(
                    "student_name",
                    ""
                ),

                class_name,

                month,

                "Not Paid",

                student.get(
                    "parent_phone",
                    ""
                ),

                ""
            ])

    # ====================================
    # AUTO COLUMN WIDTH
    # ====================================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # ====================================
    # SAVE FILE
    # ====================================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            f"fees_{month}.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


@app.route("/export-class-students")
@login_required
def export_class_students():

    class_id = request.args.get(
        "class_id"
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Students"

    headers = [

        "Student Name",

        "Parent Name",

        "Phone",

        "Class"
    ]

    ws.append(headers)

    students_ref = db.collection(
        "students"
    ).where(
        "class_id",
        "==",
        class_id
    ).stream()

    for doc in students_ref:

        s = doc.to_dict()

        class_doc = db.collection(
            "classes"
        ).document(
            s.get("class_id")
        ).get()

        class_name = ""

        if class_doc.exists:

            c = class_doc.to_dict()

            class_name = (
                f"Grade {c.get('grade')} - "
                f"{c.get('class_type')}"
            )

        ws.append([

            s.get("student_name"),

            s.get("parent_name"),

            s.get("parent_phone"),

            class_name
        ])

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            "class_students.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


# =========================================
# BROADCAST SMS
# =========================================

@app.route(
    "/broadcast",
    methods=["GET", "POST"]
)
@login_required
def broadcast():

    message = None

    # GET ALL CLASSES

    classes_ref = db.collection(
        "classes"
    ).stream()

    classes = [

        doc.to_dict()

        for doc in classes_ref
    ]

    # =====================================
    # SEND SMS
    # =====================================

    if request.method == "POST":

        class_id = request.form.get(
            "class_id"
        )

        message_type = request.form.get(
            "message_type"
        )

        custom_message = request.form.get(
            "custom_message",
            ""
        )

        # GET CLASS

        class_doc = db.collection(
            "classes"
        ).document(class_id).get()

        if not class_doc.exists:

            return render_template(

                "broadcast.html",

                classes=classes,

                message="Class not found"
            )

        class_data = class_doc.to_dict()

        grade = class_data.get(
            "grade",
            ""
        )

        class_type = class_data.get(
            "class_type",
            ""
        )

        class_name = (
            f"{grade} - {class_type}"
        )
        

        # GET STUDENTS

        students_ref = db.collection(
            "students"
        ).where(
            "class_id",
            "==",
            class_id
        ).stream()

        students = [

            doc.to_dict()

            for doc in students_ref
        ]

        sms_count = 0

        # LOOP STUDENTS

        for student in students:

            student_name = student.get(
                "student_name",
                "Student"
            )

            parent_phone = student.get(
                "parent_phone",
                ""
            )

            # =============================
            # CANCEL CLASS
            # =============================

            if message_type == "cancel":

                final_message = (

                    f"Dear Parent, "

                    f"{student_name}'s "

                    f"{class_name} "

                    f"class has been cancelled."
                )

                if custom_message:

                    final_message += (
                        f" {custom_message}"
                    )

            # =============================
            # NEW CLASS
            # =============================

            elif message_type == "new":

                final_message = (

                    f"Dear Parent, "

                    f"a new class has been scheduled for "

                    f"{student_name}. "

                    f"Class: {class_name}."
                )

                if custom_message:

                    final_message += (
                        f" {custom_message}"
                    )

            # =============================
            # GENERAL
            # =============================

            else:

                final_message = (

                    f"Dear Parent, "

                    f"{student_name} - "

                    f"{class_name}. "
                )

                if custom_message:

                    final_message += (
                        custom_message
                    )

            # =================================
            # SMS PLACE
            # =================================

            print(
                f"SMS TO {parent_phone}"
            )

            print(final_message)

            print("--------------------------------")

            # SAVE LOG

            db.collection(
                "broadcast_logs"
            ).add({

                "student_id": student.get(
                    "student_id"
                ),

                "student_name": student_name,

                "class_id": class_id,

                "class_name": class_name,

                "parent_phone": parent_phone,

                "message_type": message_type,

                "message": final_message,

                "sent_at": datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
            })

            sms_count += 1

        message = (
            f"Broadcast sent to "
            f"{sms_count} parents."
        )

    return render_template(

        "broadcast.html",

        classes=classes,

        message=message
    )


# =========================================
# TEST SMS
# =========================================

@app.route("/test-sms")
@login_required
def test_sms():

    send_sms(

        "94705240051",

        "ClassTrack SMS test successful"
    )

    return "SMS SENT"

# =========================================
# MAIN
# =========================================

if __name__ == "__main__":
    app.run(host="0.0.0.0")
